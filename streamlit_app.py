import streamlit as st
import pandas as pd
import openpyxl
import plotly.express as px

# Configuración de la página
st.set_page_config(page_title="Tablero de Indicadores", layout="wide")

st.title("📊 Tablero de Indicadores SPT")
st.write("Sube tu archivo **Tablero_Indicadores SPT.xlsx** para visualizar la asistencia y calidad de aplicación.")

# 1. Carga del archivo
archivo_subido = st.file_uploader("Sube el archivo Excel", type=['xlsx'])

if archivo_subido is not None:
    # 2. Leer el Excel y sus propiedades usando openpyxl (para los colores de las pestañas)
    wb = openpyxl.load_workbook(archivo_subido, data_only=True)
    
    nombres_participantes = []
    hojas_datos = []
    
    # Inteligencia para separar hojas por color de pestaña
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
        color_pestaña = hoja.sheet_properties.tabColor
        
        es_verde = False
        es_rojo = False
        
        # Evaluar el color RGB si existe
        if color_pestaña is not None and color_pestaña.rgb is not None:
            rgb = str(color_pestaña.rgb).upper()
            # Códigos hexadecimales comunes para verde en Excel (suelen empezar con FF seguidos de mucho verde)
            if "00B050" in rgb or "00FF00" in rgb or rgb.startswith("FF00"): 
                es_verde = True
            # Códigos comunes para rojo
            elif "FF0000" in rgb or rgb.startswith("FFFF0000"):
                es_rojo = True
        
        # Clasificación inicial basada en la detección
        if es_verde:
            nombres_participantes.append(nombre_hoja)
        else:
            # Si es rojo o no tiene color, lo mandamos a datos u otros
            hojas_datos.append(nombre_hoja)
            
    # Si la detección de color falla (por temas de Excel), ponemos todas las hojas para que el usuario elija
    if not nombres_participantes:
        nombres_participantes = wb.sheetnames

    # 3. Interfaz para confirmar la selección de participantes
    st.sidebar.header("Configuración de Datos")
    participantes_seleccionados = st.sidebar.multiselect(
        "Hojas identificadas como Participantes (Verde):",
        options=wb.sheetnames,
        default=nombres_participantes
    )
    
    if not participantes_seleccionados:
        st.warning("Por favor, selecciona al menos un participante en el menú lateral.")
    else:
        st.success("¡Archivo cargado y procesado con éxito!")
        
        # 4. Procesamiento y Visualización por Participante
        tabs = st.tabs(participantes_seleccionados)
        
        for i, participante in enumerate(participantes_seleccionados):
            with tabs[i]:
                st.subheader(f"Indicadores de: {participante}")
                
                # Leer los datos específicos de esa hoja usando pandas para mayor facilidad
                df = pd.read_excel(archivo_subido, sheet_name=participante)
                
                # --- LÓGICA DE EJEMPLO PARA GRÁFICOS ---
                # NOTA: Debes ajustar 'Asistencia' y 'Calidad' a los nombres exactos de las columnas en tu Excel.
                
                # Mostrar los datos en tabla
                st.write("**Datos registrados:**")
                st.dataframe(df, use_container_width=True)
                
                col1, col2 = st.columns(2)
                
                try:
                    # Gráfico de Asistencia (Ejemplo: Asumiendo que hay una columna 'Asistencia' con 'Sí' / 'No')
                    if 'Asistencia' in df.columns:
                        asistencia_counts = df['Asistencia'].value_counts().reset_index()
                        asistencia_counts.columns = ['Estado', 'Cantidad']
                        
                        fig_asistencia = px.pie(
                            asistencia_counts, 
                            names='Estado', 
                            values='Cantidad', 
                            title="Porcentaje de Asistencia",
                            color='Estado',
                            color_discrete_map={'Sí': '#00CC96', 'No': '#EF553B'} # Verde y Rojo
                        )
                        col1.plotly_chart(fig_asistencia, use_container_width=True)
                    else:
                        col1.info("No se encontró la columna 'Asistencia' en esta hoja para generar el gráfico.")

                    # Gráfico de Formación/Calidad (Ejemplo: Asumiendo una columna 'Formacion_Completada' o similar)
                    # Aquí usamos una métrica hipotética de % de formación
                    if '% Formación' in df.columns:
                        formacion_counts = df['% Formación'].value_counts().reset_index()
                        formacion_counts.columns = ['Nivel', 'Cantidad']
                        
                        fig_formacion = px.pie(
                            formacion_counts, 
                            names='Nivel', 
                            values='Cantidad', 
                            title="% de Formación / Calidad",
                            hole=0.4 # Estilo de dona
                        )
                        col2.plotly_chart(fig_formacion, use_container_width=True)
                    else:
                        col2.info("No se encontró la columna '% Formación' en esta hoja para generar el gráfico.")
                        
                except Exception as e:
                    st.error(f"Ocurrió un error al generar los gráficos: {e}")
